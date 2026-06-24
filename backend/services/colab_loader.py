"""Colab content loader service - fetches code from Google Colab."""
import ast
import csv
import requests
import json
import zipfile
import io
import re
from typing import Optional
import logging

from backend.config import Config
from backend.services.code_cleaner import clean_notebook_code

logger = logging.getLogger(__name__)

# Must stay in sync with the same constant in code_parser.py.
_CELL_OUTPUTS_MARKER = '# === CELL OUTPUTS ==='

# Prefixes that identify the first line of real Python code in mixed content.
_PYTHON_CODE_STARTERS = ('import ', 'from ', 'def ', 'class ', 'async ', '@')


def _can_parse_as_python(source: str) -> bool:
    """Return True when *source* can be compiled as valid Python."""
    try:
        compile(source, '<string>', 'exec', flags=ast.PyCF_ALLOW_TOP_LEVEL_AWAIT)
        return True
    except SyntaxError:
        return False


def _detect_and_mark_output_boundary(text: str) -> str:
    """Insert the cell-outputs marker at the Python-code / execution-output boundary.

    The external service returns all notebook cells (code + markdown) **plus**
    captured execution output concatenated into a single flat string.  This
    function locates where the Python code ends (i.e. the last line where the
    prefix ``lines[ps:end]`` still parses as valid Python) and injects the
    ``_CELL_OUTPUTS_MARKER`` so the downstream ``_split_code_and_trailing_output``
    fast-path can cleanly separate executable code from runtime logs.

    Strategy
    --------
    1. Find *ps* — the index of the first line that begins a Python construct
       (import, def, class, async, decorator).  Lines before *ps* are typically
       markdown headings or student-added text labels; they are kept in the
       code block but skipped during the forward scan.
    2. Scan forward from *ps* to find *last_valid* — the furthest position
       where ``lines[ps:last_valid]`` is parseable Python.
    3. Only insert the marker when the code block is a meaningful fraction of
       the total content (≥ 30 % of non-blank lines).  A very small ratio
       indicates that the scan stopped at a genuine syntax error in student
       code rather than at trailing execution output; in that case we leave the
       content unchanged so the evaluator can still see everything.
    """
    if _CELL_OUTPUTS_MARKER in text:
        return text  # already has the marker — nothing to do

    lines = text.split('\n')
    n = len(lines)

    # Step 1: locate first real-Python-code line.
    ps = None
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped and any(stripped.startswith(s) for s in _PYTHON_CODE_STARTERS):
            ps = i
            break

    if ps is None:
        return text  # no Python start found

    # Step 2: forward scan from ps.
    last_valid = ps  # exclusive end — lines[ps:last_valid] is valid Python
    for end in range(ps + 1, n + 1):
        candidate = '\n'.join(lines[ps:end])
        if _can_parse_as_python(candidate):
            last_valid = end

    if last_valid <= ps:
        return text  # couldn't build even one parseable line

    # Step 3: safeguard — skip split if code portion is too small.
    code_non_empty = sum(1 for l in lines[ps:last_valid] if l.strip())
    total_non_empty = sum(1 for l in lines if l.strip())
    if total_non_empty == 0 or code_non_empty / total_non_empty < 0.30:
        return text  # likely a genuine early syntax error — don't split

    # Step 4: check there is actually trailing content worth separating.
    if not any(l.strip() for l in lines[last_valid:]):
        return text  # nothing after the code — no marker needed

    # Insert the marker between code and trailing output.
    # Preamble lines (0 … ps-1) stay attached to the code block.
    result_lines = lines[:last_valid] + ['', _CELL_OUTPUTS_MARKER, ''] + lines[last_valid:]
    return '\n'.join(result_lines)


def _parse_service_flat_text(text: str) -> Optional[str]:
    """Parse the structured flat text returned by the external service.

    The service formats ``resheniye`` as a series of cell blocks separated by
    ``----------`` (10 dashes), each prefixed with cell-type metadata::

        =====================
        # Содержимое файла ...:
        Cell type: markdown
        Source: # ТГ-бот
        код написан ChatGPT
        ----------
        Cell type: code
        Source: import asyncio
        ...
        Execution count: 2
        Executed: True
        Executed by: ...
        Executed at: ...
        Outputs: TELEGRAM_TOKEN: OK
        ...
        ----------
        Cell type: markdown
        Source: # Ответы на вопросы
        ...
        ----------

    Extracts:
    * **code cells** → code section (joined with ``# --- Cell ---``)
    * **code-cell outputs** → placed after ``# === CELL OUTPUTS ===``
    * **markdown cells** → also placed after the outputs marker so student
      answers remain visible in the logs tab

    Returns ``None`` when the format is not recognised (triggers fallback to
    the heuristic ``_detect_and_mark_output_boundary``).
    """
    if 'Cell type:' not in text or 'Source:' not in text:
        return None

    # Split on lines that are purely dashes (8+).
    # The service uses "----------" (10 dashes) as cell separators.
    # Using 8+ avoids false splits on "---" inside Jinja2/Markdown template strings.
    blocks = re.split(r'\n-{8,}\n', text)

    code_parts: list[str] = []
    output_parts: list[str] = []

    _EXEC_META = (
        'Execution count:', 'Executed:', 'Executed by:', 'Executed at:', 'Outputs:',
    )

    for block in blocks:
        lines = block.split('\n')

        cell_type: Optional[str] = None
        source_start_idx: Optional[int] = None

        for i, line in enumerate(lines):
            stripped = line.strip()
            if stripped.startswith('Cell type:') and cell_type is None:
                cell_type = stripped[len('Cell type:'):].strip().lower()
            if stripped.startswith('Source:') and source_start_idx is None:
                source_start_idx = i

        if cell_type is None or source_start_idx is None:
            # Orphan block — no Cell type/Source header.
            # The service sometimes returns content blocks without structured headers.
            # Collect them as output so the content is visible in logs.
            orphan = block.strip()
            if orphan:
                output_parts.append(orphan)
            continue

        # First source line may have content on the same line as "Source:".
        first_line = lines[source_start_idx]
        after_prefix = re.sub(r'^[ \t]*Source:[ \t]*', '', first_line)

        # Apply the same magic-line filter to the inline Source: content.
        _first_stripped = after_prefix.strip()
        _first_is_magic = _first_stripped.startswith('!') or (
            _first_stripped.startswith('%') and not _first_stripped.startswith('%%')
        )
        source_lines: list[str] = ([] if not after_prefix.strip() or _first_is_magic else [after_prefix])
        exec_out_lines: list[str] = ([after_prefix] if after_prefix.strip() and _first_is_magic else [])
        in_outputs = False
        in_magic_continuation = _first_is_magic and _first_stripped.endswith('\\')

        # Execution artifact patterns the service embeds inline in Source sections.
        _ARTIFACTS = ('Output image saved at:', 'Output saved at:', 'Image saved at:')

        for line in lines[source_start_idx + 1:]:
            stripped = line.strip()
            if in_outputs:
                exec_out_lines.append(line)
                continue
            if any(stripped.startswith(m) for m in _EXEC_META):
                if stripped.startswith('Outputs:'):
                    in_outputs = True
                    after_outputs = stripped[len('Outputs:'):].strip()
                    if after_outputs:
                        exec_out_lines.append(after_outputs)
                in_magic_continuation = False
                continue  # skip metadata lines
            # Execution artifacts (e.g. matplotlib image paths) placed by the
            # service inside Source — treat them as output, not code.
            if any(stripped.startswith(a) for a in _ARTIFACTS):
                exec_out_lines.append(line)
                in_magic_continuation = False
                continue
            # IPython magic lines (!pip install, %matplotlib, etc.) are Jupyter
            # shell/magic commands — not valid Python.  Route to output so they
            # are preserved in logs but excluded from executable code.
            # Also handle multi-line continuations (lines ending with \).
            is_magic_start = stripped.startswith('!') or (
                stripped.startswith('%') and not stripped.startswith('%%')
            )
            if is_magic_start or in_magic_continuation:
                exec_out_lines.append(line)
                # Continue consuming continuation lines until no trailing backslash.
                in_magic_continuation = stripped.endswith('\\')
                continue
            source_lines.append(line)

        source = '\n'.join(source_lines).rstrip()

        if cell_type == 'code':
            if source.strip():
                code_parts.append(source)
            if exec_out_lines:
                output_parts.extend(exec_out_lines)
        else:
            # markdown / raw cells → put content in the output section so
            # student answers remain visible to the evaluator in the logs tab.
            if source.strip():
                output_parts.append(source.strip())

    if not code_parts:
        # Text-only notebook (zero code cells) — all content goes to the
        # outputs section so downstream code_parser puts it in logs, not code.
        if output_parts:
            return _CELL_OUTPUTS_MARKER + '\n' + '\n'.join(output_parts)
        return None

    result = '\n\n# --- Cell ---\n\n'.join(code_parts)
    if output_parts:
        result += '\n\n' + _CELL_OUTPUTS_MARKER + '\n' + '\n'.join(output_parts)
    return result



    """
    Extract Google Drive file ID from a URL.
    
    Args:
        url: Google Drive URL
        
    Returns:
        File ID or None
    """
    # Try /d/{id} pattern
    match = re.search(r'/d/([a-zA-Z0-9_-]+)', url)
    if match:
        return match.group(1)
    
    # Try ?id={id} pattern
    match = re.search(r'[?&]id=([a-zA-Z0-9_-]+)', url)
    if match:
        return match.group(1)
    
    return None


def _decode_buffer(buf: bytes) -> str:
    """
    Try to decode buffer as text with fallback for Russian encodings.
    
    Args:
        buf: Bytes to decode
        
    Returns:
        Decoded string
    """
    try:
        text = buf.decode('utf-8')
        # Check for replacement character (indicator of decode failure)
        if '\uFFFD' in text:
            # Try Windows-1251
            return buf.decode('cp1251', errors='replace')
        return text
    except (UnicodeDecodeError, AttributeError):
        try:
            return buf.decode('cp1251', errors='replace')
        except:
            return ""


def _is_binary_content(buf: bytes) -> bool:
    """Heuristic check for binary/non-text file content."""
    if not buf:
        return False

    # If it decodes cleanly as UTF-8 it is definitely text (e.g. .ipynb with
    # Cyrillic/CJK content whose high bytes would otherwise trip the ratio check).
    try:
        buf.decode('utf-8')
        return False
    except UnicodeDecodeError:
        pass

    # Null bytes are a strong binary indicator.
    if b'\x00' in buf:
        return True

    sample = buf[:4096]
    if not sample:
        return False

    # Count non-printable bytes excluding common whitespace.
    text_like = set(range(32, 127)) | {9, 10, 13}
    non_text = sum(1 for b in sample if b not in text_like)
    ratio = non_text / len(sample)

    # Keep threshold conservative to avoid false positives.
    return ratio > 0.35


def _extract_ipynb_code(content: str) -> str:
    """
    Extract Python code from Jupyter notebook JSON.
    
    Args:
        content: Notebook JSON content
        
    Returns:
        Concatenated code from all code cells, followed by a
        '# === CELL OUTPUTS ===' section that contains the captured
        cell execution outputs (stdout/stderr/results).
    """
    # Check if this is HTML (error page) not JSON
    if content.strip().startswith('<!DOCTYPE') or content.strip().startswith('<html'):
        raise ValueError("Received HTML instead of notebook content (likely auth error or wrong URL)")
    
    try:
        notebook = json.loads(content)
        cells = notebook.get('cells', [])
        code_cells = [c for c in cells if c.get('cell_type') == 'code']
        
        if not code_cells:
            # Valid notebook with only markdown/images and no code cells.
            return ""
        
        code_parts = []
        output_parts = []

        for cell in code_cells:
            source = cell.get('source', [])
            if isinstance(source, list):
                code_parts.append(''.join(source))
            elif isinstance(source, str):
                code_parts.append(source)

            # Capture cell execution outputs (stdout, stderr, rich results).
            for out in cell.get('outputs', []):
                out_type = out.get('output_type', '')
                if out_type in ('stream', 'execute_result', 'display_data'):
                    text = out.get('text', [])
                    if isinstance(text, list):
                        text = ''.join(text)
                    if isinstance(text, str) and text.strip():
                        output_parts.append(text.strip())
        
        result = '\n\n# --- Cell ---\n\n'.join(code_parts)

        if output_parts:
            result += '\n\n# === CELL OUTPUTS ===\n' + '\n'.join(output_parts)
        
        # Final check: extracted code should not look like HTML
        if result.strip().startswith('<') or '<!DOCTYPE' in result:
            raise ValueError("Extracted content appears to be HTML, not Python code")
        
        return result
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON in notebook: {str(e)}")


def extract_file_id_from_drive_url(url: str) -> Optional[str]:
    """Extract Google Drive file ID from various URL formats."""
    patterns = [
        r'/file/d/([a-zA-Z0-9_-]+)',
        r'/drive/([a-zA-Z0-9_-]+)',   # colab.research.google.com/drive/{id}
        r'id=([a-zA-Z0-9_-]+)',
        r'/d/([a-zA-Z0-9_-]+)',
    ]
    for pattern in patterns:
        m = re.search(pattern, url)
        if m:
            return m.group(1)
    return None


def _fetch_google_doc_as_text(url: str) -> str:
    """Fetch a Google Docs document as plain text via the export API.

    Works for any publicly-shared Google Docs link — no auth required.
    The export endpoint returns clean UTF-8 plain text regardless of
    how complex the document formatting is.
    """
    m = re.search(r'/d/([a-zA-Z0-9_-]+)', url)
    if not m:
        raise Exception("Could not extract document ID from Google Docs URL")
    doc_id = m.group(1)
    export_url = f"https://docs.google.com/document/d/{doc_id}/export?format=txt"
    logger.info(f"Google Docs export: {export_url}")
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
    r = requests.get(export_url, headers=headers, timeout=25, allow_redirects=True)
    if r.status_code != 200:
        raise Exception(f"Google Docs export failed: HTTP {r.status_code}")
    text = r.content.decode('utf-8', errors='replace')
    if not text.strip():
        raise Exception("Google Docs export returned empty content")
    return _CELL_OUTPUTS_MARKER + '\n' + text


def _csv_to_html_table(raw_csv: str, sheet_name: str = '') -> str:
    """Convert a CSV string to an HTML table, stripping trailing empty columns."""
    rows = list(csv.reader(raw_csv.splitlines()))
    if not rows:
        raise Exception("Google Sheets: no data rows found")

    # Find the maximum non-empty column index across all rows so we can
    # drop trailing all-empty columns that Google sometimes pads in gviz output.
    max_col = 0
    for row in rows:
        for i in range(len(row) - 1, -1, -1):
            if row[i].strip():
                max_col = max(max_col, i)
                break

    lines = []
    if sheet_name:
        lines.append(f'<h2>Вкладка: {sheet_name}</h2>')
    lines.append('<table border="1" cellpadding="4" cellspacing="0">')
    for i, row in enumerate(rows):
        trimmed = row[:max_col + 1]
        lines.append('<tr>')
        tag = 'th' if i == 0 else 'td'
        for cell in trimmed:
            escaped = cell.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            lines.append(f'  <{tag}>{escaped}</{tag}>')
        lines.append('</tr>')
    lines.append('</table>')
    return '\n'.join(lines)


# Colors that are equivalent to «no fill» or «default black text» in XLSX.
_XLSX_NO_COLOR = frozenset({
    '00000000',  # fully transparent
    'FF000000',  # opaque black (default font)
    '000000',    # same without alpha
    'FFFFFFFF',  # opaque white (default background)
    'FFFFFF',
})


def _xlsx_color_to_css(color_obj) -> Optional[str]:
    """Return a CSS hex color string from an openpyxl Color, or None if it is a default."""
    if color_obj is None:
        return None
    if getattr(color_obj, 'type', None) != 'rgb':
        return None
    rgb: str = getattr(color_obj, 'rgb', '') or ''
    if not rgb or rgb.upper() in _XLSX_NO_COLOR:
        return None
    # Strip the alpha channel byte (openpyxl uses 8-char ARGB).
    return f'#{rgb[2:] if len(rgb) == 8 else rgb}'


def _xlsx_sheets_to_html(content: bytes) -> Optional[str]:
    """Parse XLSX bytes and return all sheets as HTML tables with inline cell colors.

    Returns None when openpyxl is unavailable or the file cannot be parsed.
    Each sheet becomes an ``<h2>`` + ``<table>`` block.  Cell background and
    font colors are encoded as inline ``style`` attributes so the LLM evaluator
    can reason about colour-coded answers.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed — falling back to CSV export")
        return None

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        logger.warning(f"openpyxl failed to open workbook: {e}")
        return None

    sheet_parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_column is None:
            continue

        # Find the last column that has any non-empty cell value in the sheet.
        used_max_col = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in reversed(row):
                if cell.value is not None and str(cell.value).strip():
                    used_max_col = max(used_max_col, cell.column)
                    break
        if used_max_col == 0:
            continue

        rows_html: list[str] = []
        any_content = False

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=used_max_col)
        ):
            cells_html: list[str] = []
            row_has_content = False

            for cell in row:
                raw = cell.value
                # Format numbers cleanly: integers without decimal point.
                if isinstance(raw, float) and raw == int(raw):
                    value = str(int(raw))
                elif raw is None:
                    value = ''
                else:
                    value = str(raw)
                if value.strip():
                    row_has_content = True

                style_parts: list[str] = []

                # Background fill color.
                fill = getattr(cell, 'fill', None)
                if fill and getattr(fill, 'fill_type', None) == 'solid':
                    bg = _xlsx_color_to_css(getattr(fill, 'fgColor', None))
                    if bg:
                        style_parts.append(f'background-color: {bg}')

                # Font color.
                font = getattr(cell, 'font', None)
                if font:
                    fc = _xlsx_color_to_css(getattr(font, 'color', None))
                    if fc:
                        style_parts.append(f'color: {fc}')

                style_attr = f' style="{"; ".join(style_parts)}"' if style_parts else ''
                escaped = value.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                tag = 'th' if row_idx == 0 else 'td'
                cells_html.append(f'  <{tag}{style_attr}>{escaped}</{tag}>')

            if row_has_content or row_idx == 0:
                rows_html.append('<tr>\n' + '\n'.join(cells_html) + '\n</tr>')
                any_content = True

        if not any_content:
            continue

        escaped_name = sheet_name.replace('&', '&amp;').replace('<', '&lt;')
        sheet_parts.append(
            f'<h2>Вкладка: {escaped_name}</h2>\n'
            f'<table border="1" cellpadding="4" cellspacing="0">\n'
            + '\n'.join(rows_html)
            + '\n</table>'
        )

    if not sheet_parts:
        return None

    logger.info(
        f"Google Sheets XLSX: {len(sheet_parts)} sheet(s) parsed → "
        + ', '.join(f'«{n}»' for n in wb.sheetnames)
    )
    return '\n\n'.join(sheet_parts)


def _fetch_sheet_csv_via_gviz(doc_id: str, gid: str, req_headers: dict) -> str:
    """Fetch a single sheet as CSV via the gviz/tq endpoint."""
    gviz_url = f"https://docs.google.com/spreadsheets/d/{doc_id}/gviz/tq?tqx=out:csv&gid={gid}"
    logger.info(f"Google Sheets gviz CSV: {gviz_url}")
    r = requests.get(gviz_url, headers=req_headers, timeout=25, allow_redirects=True)
    if r.status_code == 200 and r.content:
        candidate = r.content.decode('utf-8', errors='replace')
        if not candidate.strip().startswith('<'):
            return candidate
    raise Exception(
        f"Google Sheets gviz export failed (HTTP {r.status_code}). "
        "Make sure the spreadsheet is shared as «Anyone with the link can view»."
    )


def _fetch_google_sheet_as_html(url: str) -> str:
    """Fetch all sheets of a Google Spreadsheet as HTML tables with cell colors.

    Priority:
    1. XLSX export → openpyxl: reads every sheet tab and preserves background /
       font colours as inline ``style`` attributes.
    2. gviz/tq CSV: text-only fallback for the sheet indicated by ``gid`` in the
       URL when XLSX export is unavailable or openpyxl is not installed.
    """
    m = re.search(r'/d/([a-zA-Z0-9_-]+)', url)
    if not m:
        raise Exception("Could not extract spreadsheet ID from Google Sheets URL")
    doc_id = m.group(1)

    gid_match = re.search(r'[?&#]gid=(\d+)', url)
    gid = gid_match.group(1) if gid_match else '0'

    req_headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

    # PRIMARY: XLSX export (all sheets + cell colors via openpyxl).
    xlsx_url = f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=xlsx"
    logger.info(f"Google Sheets XLSX export: {xlsx_url}")
    try:
        r = requests.get(xlsx_url, headers=req_headers, timeout=30, allow_redirects=True)
        # XLSX is a ZIP — magic bytes PK (0x50 0x4B).
        if r.status_code == 200 and len(r.content) > 100 and r.content[:2] == b'PK':
            html = _xlsx_sheets_to_html(r.content)
            if html:
                return _CELL_OUTPUTS_MARKER + '\n' + html
            logger.warning("XLSX parsed but produced no content — falling back to gviz CSV")
        else:
            logger.warning(
                f"Google Sheets XLSX export: HTTP {r.status_code} — falling back to gviz CSV"
            )
    except Exception as e:
        logger.warning(f"Google Sheets XLSX export error: {e} — falling back to gviz CSV")

    # FALLBACK: gviz/tq CSV (text only, single sheet from URL gid).
    raw_csv = _fetch_sheet_csv_via_gviz(doc_id, gid, req_headers)
    table_html = _csv_to_html_table(raw_csv)
    logger.info("Google Sheets: gviz CSV fallback used (no colors, single sheet)")
    return _CELL_OUTPUTS_MARKER + '\n' + table_html


def _try_parse_docx(content: bytes) -> Optional[str]:
    """Try to parse bytes as a .docx (Office Open XML) file.

    Returns the extracted plain text wrapped in the cell-outputs marker,
    or ``None`` if the content is not a valid .docx or python-docx is not installed.
    """
    try:
        from docx import Document  # python-docx
        doc = Document(io.BytesIO(content))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        if not paragraphs:
            return None
        text = '\n'.join(paragraphs)
        return _CELL_OUTPUTS_MARKER + '\n' + text
    except Exception:
        return None


def get_colab_content(url: str, assignment_type: str = "code") -> str:
    """
    Fetch code content from Google Colab or archive.

    Priority order:
    1. Google Sheets export API (if docs.google.com/spreadsheets) → HTML table
    2. Google Docs export API (if docs.google.com/document) → plain text
    3. External service (primary for Colab/Drive — parses by cell type)
    4. Direct Drive download (fallback)
    5. drive.usercontent.google.com (last-resort fallback)
    """
    logger.info(f"Fetching Colab content: {url}")

    # Google Sheets — always use CSV export API (viewer page is HTML only)
    if 'docs.google.com/spreadsheets' in url:
        return _fetch_google_sheet_as_html(url)

    # Google Docs — always use plain-text export API (viewer page is HTML only)
    if 'docs.google.com/document' in url:
        return _fetch_google_doc_as_text(url)

    original_url = url  # preserved for external service call
    file_id = ""
    service_error = "Not attempted"
    last_error = "Not attempted"

    # Extract file_id for use in direct-download fallbacks
    if 'colab.research.google.com' in url or 'drive.google.com' in url:
        file_id = extract_file_id_from_drive_url(url)
        if file_id:
            logger.info(f"Extracted file_id: {file_id}")
        else:
            logger.warning("Could not extract file ID from Drive/Colab URL")

    # ── PRIMARY: External service (parses by cell type — no heuristics) ──
    try:
        logger.info("Trying external Colab service...")
        service_response = requests.post(
            Config.EXTERNAL_SERVICE_URL,
            json={"url_solving": original_url},
            headers={
                "Content-Type": "application/json",
                "Authorization": Config.EXTERNAL_SERVICE_AUTH
            },
            timeout=15
        )

        data = service_response.json()
        if data.get('state_load'):
            logger.info("External service success")
            content = data.get('resheniye', '')
            if content.strip().startswith('{') and '"cells"' in content:
                logger.info("External service returned .ipynb JSON — extracting code cells")
                content = _extract_ipynb_code(content)
                return content  # cleanly extracted by cell type — no further cleaning
            else:
                parsed = _parse_service_flat_text(content)
                if parsed is not None:
                    logger.info("External service: parsed structured flat text (Cell type markers found)")
                    return parsed  # cleanly extracted by cell type — no further cleaning
                else:
                    logger.info("External service: heuristic boundary detection (no Cell type markers)")
                    content = _detect_and_mark_output_boundary(content)
                    return clean_notebook_code(content)  # dirty — clean it
        else:
            service_error = data.get('error', 'Unknown error')
            raise Exception(f"Service failed: {service_error}")

    except Exception as e:
        service_error = str(e)
        logger.warning(f"External service failed: {service_error}. Trying direct fetch...")

    # ── FALLBACK: Direct Drive/URL download ──
    target_url = url
    if file_id:
        target_url = f"https://drive.google.com/uc?id={file_id}&export=download&confirm=t"
        logger.info(f"Direct fallback URL: {target_url}")

    try:
        logger.info(f"Attempting direct fetch: {target_url}")

        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }

        response = requests.get(
            target_url,
            headers=headers,
            timeout=25,
            allow_redirects=True,
            stream=True
        )

        # Handle Google Drive virus scan warning
        if response.status_code == 200:
            content = response.content
            content_str = content.decode('utf-8', errors='ignore')

            if 'confirm=' in content_str and file_id:
                confirm_match = re.search(r'confirm=([a-zA-Z0-9_-]+)', content_str)
                if confirm_match:
                    confirm_token = confirm_match.group(1)
                    cookies = response.headers.get('set-cookie', '')
                    confirmed_url = f"https://drive.google.com/uc?export=download&confirm={confirm_token}&id={file_id}"
                    logger.info(f"Following Drive confirmation: {confirmed_url}")
                    response = requests.get(
                        confirmed_url,
                        headers={**headers, 'Cookie': cookies},
                        timeout=25,
                        allow_redirects=True
                    )
                    content = response.content

        if response.status_code != 200:
            raise Exception(f"HTTP {response.status_code}: {response.reason}")

        content_type = response.headers.get('content-type', '')
        logger.info(f"Direct fetch success: Content-Type={content_type}, size={len(content)}")

        is_zip = (
            'application/zip' in content_type or
            target_url.lower().endswith('.zip') or
            (len(content) > 4 and content[0:2] == b'PK')
        )

        # Binary payloads (images, PDFs, etc.)
        if (
            content_type.startswith('image/') or
            'application/pdf' in content_type or
            _is_binary_content(content)
        ):
            if assignment_type == "text":
                docx_text = _try_parse_docx(content)
                if docx_text:
                    logger.info("Detected .docx for text assignment; extracted text")
                    return docx_text
            logger.info("Detected binary file content; returning empty code")
            return ""

        if is_zip:
            logger.info("Detected ZIP archive, unzipping...")
            try:
                with zipfile.ZipFile(io.BytesIO(content)) as zip_file:
                    entries = zip_file.namelist()
                    logger.info(f"ZIP contents: {len(entries)} entries")

                    candidates = [
                        e for e in entries
                        if not e.endswith('/') and
                        '__MACOSX' not in e and
                        '.DS_Store' not in e and
                        (e.lower().endswith('.ipynb') or e.lower().endswith('.py') or e.lower().endswith('.txt'))
                    ]

                    if not candidates:
                        if assignment_type == "text":
                            docx_text = _try_parse_docx(content)
                            if docx_text:
                                logger.info("Detected .docx; extracted text content")
                                return docx_text
                        logger.info("ZIP has no .ipynb/.py/.txt files; returning empty code")
                        return ""

                    candidates.sort(key=lambda e: (
                        not e.lower().endswith('.ipynb'),
                        not e.lower().endswith('.py'),
                        -zip_file.getinfo(e).file_size
                    ))
                    target_entry = candidates[0]
                    logger.info(f"Selected: {target_entry}")

                    file_content = zip_file.read(target_entry)
                    decoded = _decode_buffer(file_content)
                    if target_entry.lower().endswith('.ipynb'):
                        decoded = _extract_ipynb_code(decoded)
                    return clean_notebook_code(decoded)

            except zipfile.BadZipFile as e:
                last_error = f"Bad ZIP: {str(e)}"
                logger.warning(f"ZIP error: {last_error}")
        else:
            decoded = _decode_buffer(content)

            if decoded.strip().startswith('<') or '<!DOCTYPE' in decoded or '<html' in decoded:
                raise Exception("Received HTML page instead of file content")

            if decoded.strip().startswith('{') and '"cells"' in decoded:
                logger.info("Detected .ipynb content")
                decoded = _extract_ipynb_code(decoded)
                if not decoded.strip():
                    last_error = "Notebook code extraction returned no cells"
                    logger.warning(last_error)
                else:
                    return decoded  # cleanly extracted by cell type — no further cleaning

            if len(decoded.strip()) > 10:
                if assignment_type == "text" and _CELL_OUTPUTS_MARKER not in decoded:
                    logger.info("Text assignment: wrapping plain content as logs")
                    return _CELL_OUTPUTS_MARKER + '\n' + decoded
                return clean_notebook_code(decoded)  # dirty text — clean it
            else:
                last_error = "Downloaded content too small or empty"

    except Exception as e:
        last_error = str(e)
        logger.warning(f"Direct fetch failed: {last_error}")

    # ── LAST RESORT: drive.usercontent.google.com ──
    if file_id:
        try:
            alt_url = f"https://drive.usercontent.google.com/download?id={file_id}&export=download&confirm=t&authuser=0"
            logger.info(f"Trying usercontent fallback: {alt_url}")
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
            alt_response = requests.get(alt_url, headers=headers, timeout=25, allow_redirects=True)
            if alt_response.status_code == 200:
                alt_decoded = _decode_buffer(alt_response.content)
                if not (alt_decoded.strip().startswith('<') or '<!DOCTYPE' in alt_decoded or '<html' in alt_decoded.lower()[:200]):
                    if alt_decoded.strip().startswith('{') and '"cells"' in alt_decoded:
                        logger.info("usercontent fallback: detected .ipynb content")
                        alt_decoded = _extract_ipynb_code(alt_decoded)
                    if len(alt_decoded.strip()) > 10:
                        logger.info("usercontent fallback succeeded")
                        return clean_notebook_code(alt_decoded)
        except Exception as alt_e:
            logger.warning(f"usercontent fallback failed: {alt_e}")

    raise Exception(
        f"Failed to load content. Service: {service_error}. Direct: {last_error}"
    )
